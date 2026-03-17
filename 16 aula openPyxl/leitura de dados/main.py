from openpyxl import load_workbook

arquivo = load_workbook('16 aula openPyxl/leitura de dados/alunos.xlsx')
planilha_alunos = arquivo['Alunos']

Dados1 = planilha_alunos['B2'].value
Dados2 = planilha_alunos['D5'].value
Dados3 = planilha_alunos['E10'].value


print(Dados1,"/", Dados2,"/", Dados3)

for celula in planilha_alunos['D']:
    if celula.row == 1:
        continue
    if celula.value > 8:
        print(celula.value)

for linha in planilha_alunos.iter_rows(values_only=True, min_row=2):
    print('-'*50)
    nome, Curso, idade, Nota_Final, Data_Matricula = linha
    print(f"""NOME: {nome}
CURSO: {Curso}
IDADE: {idade}
Nota: {Nota_Final}
DATA DE MATRÍCULA: {Data_Matricula.strftime("%d/%m/%Y")}""")