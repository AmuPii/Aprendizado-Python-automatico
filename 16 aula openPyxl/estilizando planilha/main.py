from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import datetime

arquivo = Workbook()
planilha = arquivo.active
planilha.title = 'Produtos'

config_fonte_cabecalho = Font(bold=True, color='00FFFF')
config_fundo_cabecalho = PatternFill(fgColor="4f4f4f", fill_type="solid")

planilha.merge_cells('A1:D1')
planilha['A1'] = 'LISTA DE PRODUTOS'
planilha.append(['Produto', 'Preço', 'Quantidade', 'Data'])

for celula in planilha[2]:
    celula.font = config_fonte_cabecalho
    celula.fill = config_fundo_cabecalho

Agr = datetime.datetime.now()

planilha.append(['Camiseta', 59.99, 10, Agr])
planilha.append(['Calça Jeans', 120.00, 5, Agr])
planilha.append(['Tênis', 250.00, 2, Agr])

# Configurações de alinhamento
for celula in planilha['B']:
    if celula.row <= 2:
        continue
    celula.alignment = Alignment(horizontal='center')
    celula.number_format = 'R$ #,##0.00'

# Configurações de data
for celula in planilha['D']:
    celula.number_format = 'DD/MM/YY'

# Configurações de borda
fina = Side(style='thin')
borda = Border(left=fina, right=fina, top=fina, bottom=fina)
for linha in planilha.iter_rows():
    for celula in linha:
        celula.border = borda



arquivo.save('16 aula openPyxl/estilizando planilha/produtos.xlsx')