from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


arquivo = Workbook()
planilha = arquivo.active
planilha.title = 'Diário de Leituras'

planilha.merge_cells('A1:D1')
planilha['A1'] = 'Diário de Leituras – Agosto 2025'

planilha['A1'].alignment = Alignment(horizontal='center')
planilha['A1'].fill = PatternFill(fgColor="1F497D", fill_type="solid")
planilha['A1'].font = Font(bold=True, color='FFFFFF', size=14)

cabecalho = ['Livro', 'Autor', 'Data de Início', 'Progresso (%)']
planilha.append(cabecalho)

header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(fgColor="4f4f4f", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col in range(1, 5):
    cell = planilha.cell(row=2, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

dados = [
    ["Sombras de Aethel", "Mariana V. Duarte", "12/01/2026", 0.45],
    ["O Código de Orion", "Lucas Ferraz", "03/02/2026", 0.78],
    ["Fragmentos do Infinito", "Rafael Mendes", "25/12/2025", 0.60],
    ["A Última Aurora", "Camila Nogueira", "10/03/2026", 0.22],
    ["Ecos do Passado", "Henrique S. Lima", "18/01/2026", 0.90],
    ["Cidade Submersa", "Beatriz Carvalho", "05/04/2026", 0.15],
    ["O Jardim das Ilusões", "Felipe Andrade", "27/02/2026", 0.66],
    ["Crônicas de Velkar", "Eduardo Tavares", "14/03/2026", 0.33],
    ["Além do Horizonte Cinzento", "Juliana Pires", "01/04/2026", 0.10],
    ["Labirinto de Cinzas", "Gustavo Rocha", "22/01/2026", 0.85]
]

for linha in dados:
    planilha.append(linha)

branco = PatternFill(fgColor="FFFFFF", fill_type="solid")
cinza_claro = PatternFill(fgColor="F2F2F2", fill_type="solid")

for row in range(3, 3 + len(dados)):
    
    fill = branco if (row % 2 == 1) else cinza_claro
    
    for col in range(1, 5):
        cell = planilha.cell(row=row, column=col)
        cell.fill = fill
        cell.border = thin_border
        
        
        if col == 1 or col == 2:                   
            cell.alignment = Alignment(horizontal='left', vertical='center')
        elif col == 3:                              
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = 'DD/MM/YYYY'
        elif col == 4:                              
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.number_format = '0.0%'

# Ajustar largura das colunas (melhor visualização)
larguras = [40, 25, 15, 15]
for i, largura in enumerate(larguras, start=1):
    planilha.column_dimensions[get_column_letter(i)].width = largura


arquivo.save('16 aula openPyxl/exercicio/diario_leituras.xlsx')