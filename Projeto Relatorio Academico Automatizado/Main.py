from openpyxl import load_workbook, Workbook

# ====================== CONFIGURAÇÃO ======================
ARQUIVO_ORIGINAL = 'Projeto Relatorio Academico Automatizado/alunos.xlsx'
ARQUIVO_APROVADOS = 'Projeto Relatorio Academico Automatizado/aprovados.xlsx'
ARQUIVO_REPROVADOS = 'Projeto Relatorio Academico Automatizado/reprovados.xlsx'

# Carregar a planilha existente
wb = load_workbook(ARQUIVO_ORIGINAL, data_only=True)  # data_only=True para pegar valores reais (não fórmulas)
ws = wb.active

# ====================== ENCONTRAR COLUNAS ======================
# Pega o cabeçalho da primeira linha
header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

nome_col_idx = None
nota_final_col_idx = None

for col_idx, valor in enumerate(header_row, start=1):
    if valor is not None:
        texto = str(valor).strip().lower()
        if texto == 'nome':
            nome_col_idx = col_idx
        elif texto in ['nota final', 'notafinal', 'nota_final']:
            nota_final_col_idx = col_idx

if nome_col_idx is None or nota_final_col_idx is None:
    raise ValueError(
        "Não foram encontradas as colunas 'Nome' e/ou 'Nota Final' no cabeçalho.\n"
        "Verifique se o cabeçalho da planilha contém exatamente essas palavras "
        "(pode ser maiúsculo/minúsculo ou com espaços)."
    )

# ====================== PROCESSAMENTO DOS ALUNOS ======================
aprovados = []
reprovados = []

soma_notas = 0.0
total_alunos = 0
maior_nota = -1.0
aluno_maior_nota = ""

for linha in ws.iter_rows(min_row=2, values_only=True):
    # Pula linhas vazias
    if not linha or linha[0] is None:
        continue

    nome = linha[nome_col_idx - 1]
    nota_raw = linha[nota_final_col_idx - 1]

    # Converte nota para float (aceita número ou texto "7.5")
    try:
        nota = float(nota_raw)
    except (ValueError, TypeError):
        continue  # ignora linha se a nota não for válida

    # Acumula dados para estatísticas
    soma_notas += nota
    total_alunos += 1

    # Aluno com maior nota
    if nota > maior_nota:
        maior_nota = nota
        aluno_maior_nota = nome if nome is not None else "Desconhecido"

    # Separa em grupos
    if nota >= 7.0:
        aprovados.append(linha)
    else:
        reprovados.append(linha)

# ====================== CÁLCULO DA MÉDIA ======================
if total_alunos > 0:
    media_turma = soma_notas / total_alunos
else:
    media_turma = 0.0
    aluno_maior_nota = "Nenhum aluno encontrado"

# ====================== CRIAR NOVOS ARQUIVOS ======================
# Cabeçalho original (para copiar nas duas planilhas)
cabecalho = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

# --- Aprovados ---
wb_aprovados = Workbook()
ws_aprovados = wb_aprovados.active
ws_aprovados.title = 'Aprovados'
ws_aprovados.append(cabecalho)          # copia cabeçalho
for linha in aprovados:
    ws_aprovados.append(linha)
wb_aprovados.save(ARQUIVO_APROVADOS)

# --- Reprovados ---
wb_reprovados = Workbook()
ws_reprovados = wb_reprovados.active
ws_reprovados.title = 'Reprovados'
ws_reprovados.append(cabecalho)         # copia cabeçalho
for linha in reprovados:
    ws_reprovados.append(linha)
wb_reprovados.save(ARQUIVO_REPROVADOS)

# ====================== EXIBIR RESULTADOS NO TERMINAL ======================
print("=" * 50)
print("✅ PROCESSAMENTO CONCLUÍDO!")
print("=" * 50)
print(f"Quantidade de aprovados: {len(aprovados)}")
print(f"Quantidade de reprovados: {len(reprovados)}")
print(f"Nota média da turma: {media_turma:.2f}")
print(f"Nome do aluno com a maior nota: {aluno_maior_nota} ({maior_nota:.1f})")
print("=" * 50)
print(f"📁 Arquivos criados com sucesso:")
print(f"   • {ARQUIVO_APROVADOS}")
print(f"   • {ARQUIVO_REPROVADOS}")
print("=" * 50)